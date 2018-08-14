#!/usr/bin/python
# codding:utf-8
import os
import glob
import openpyxl


def merge_xlsx_file(filenames):
	if len(filenames)<2:
		return None
	wb = openpyxl.load_workbook(filenames[0])
	ws = wb.active
	ws.title = 'mergeed result'
	
	for filename in filenames[1:]:
		workbook = openpyxl.load_workbook(filename)
		sheet = workbook.active
		for row in sheet.iter_rows(min_row=3,min_col=1):
			values = [cell.value for cell in row]
			ws.append(values)
	return wb

def find_all_xlsx(path):
	sign = '*.xlsx'
	files = glob.glob(os.path.join(path,sign))
	sorted(files,key=str.lower)
	return files


def main():
	files = find_all_xlsx('./')
	wb = merge_xlsx_file(files)
	if wb:
		wb.save('merge_file.xlsx')
	else:
		print 'not file merge'

if __name__ == '__main__':
	main() 
